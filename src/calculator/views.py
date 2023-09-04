import requests
import openpyxl
from openpyxl.styles import Alignment
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.conf import settings
from django.forms.models import model_to_dict
from .forms import ImperialForm, MetricForm
from authentication.forms import SignUp
from calculator.models import UserStat, Diet, MacroDistribution


def round_data(data):
    for key, value in data.items():
        if isinstance(value, (float)):
            data[key] = int(round(value))
        elif isinstance(value, dict):
            round_data(value)


def retrieve_macros(form_data):
    url = "https://fitness-calculator.p.rapidapi.com/macrocalculator"
    headers = {
        "X-RapidAPI-Key": settings.API_KEY,
        "X-RapidAPI-Host": "fitness-calculator.p.rapidapi.com",
    }
    response = requests.get(url, headers=headers, params=form_data)
    response_data = response.json()["data"]

    round_data(response_data)

    return response_data


def create_spreadsheet(request, pk):
    template = "calculator/templates/calculator/Template.xlsx"
    workbook = openpyxl.load_workbook(template)
    sheet = workbook["Stats"]

    user_stat = UserStat.objects.get(pk=pk)
    stats = model_to_dict(user_stat)

    sheet.append(
        [
            user_stat.last_updated.strftime("%m/%d/%Y"),
            stats["age"],
            stats["sex"],
            f'{stats["weight_kg"]} kg',
            f'{stats["height_cm"]} cm',
            f'{stats["weight_lb"]} lb',
            f'{stats["height_ft"]}\'{stats["height_in"]}"',
            int(stats["activity_level"]),
            stats["weight_goal"],
        ]
    )

    for row in sheet.iter_rows(min_row=2, max_row=2, min_col=2, max_col=9):
        for cell in row:
            cell.alignment = Alignment(horizontal="left")

    for plan in user_stat.diet.macro_distribution.all():
        sheet = workbook[plan.plan_name]

        sheet.cell(row=3, column=1, value=plan.user_diet.calorie)
        sheet.cell(row=3, column=2, value=plan.protein)
        sheet.cell(row=3, column=3, value=plan.carbs)
        sheet.cell(row=3, column=4, value=plan.fat)

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Tracker.xlsx'

    workbook.save(response)

    return response


def calculate_macros(request, form_choice):
    if request.method == "POST":
        user = request.user
        authenticated = user.is_authenticated
        stat_instance = None

        if authenticated:
            try:
                stat_instance = UserStat.objects.get(user=request.user)
            except UserStat.DoesNotExist:
                pass

        form = (
            form_choice(request.POST, instance=stat_instance)
            if stat_instance and authenticated
            else form_choice(request.POST)
        )

        if form.is_valid():
            form_data = {
                "age": form.cleaned_data["age"],
                "gender": form.cleaned_data["sex"],
                "weight": form.cleaned_data["weight_kg"],
                "height": form.cleaned_data["height_cm"],
                "activitylevel": form.cleaned_data["activity_level"],
                "goal": form.cleaned_data["weight_goal"],
            }

            try:
                response_data = retrieve_macros(form_data)

                balanced_data = response_data["balanced"]
                lowcarb_data = response_data["lowcarbs"]
                lowfat_data = response_data["lowfat"]
                highprotein_data = response_data["highprotein"]

                if not authenticated:
                    form_type = request.POST.get("form_type")

                    if form_type == "metric":
                        imperial_form = ImperialForm()
                        metric_form = form
                    elif form_type == "imperial":
                        imperial_form = form
                        metric_form = MetricForm()
                    else:
                        metric_form = MetricForm()
                        imperial_form = ImperialForm()

                    context = {
                        "form_type": form_type,
                        "macros": response_data,
                        "metric_form": metric_form,
                        "imperial_form": imperial_form,
                    }

                    return render(request, "calculator/profile.html", context)

                elif authenticated and stat_instance:
                    stat_instance = form.save()
                    diet = Diet.objects.get(stats=stat_instance)
                    diet.calorie = response_data["calorie"]
                    diet.save(update_fields=["calorie"])

                    macro_names = ["balanced", "lowcarbs", "lowfat", "highprotein"]
                    macro_distribution = MacroDistribution.objects.filter(
                        user_diet=diet
                    )

                    for i, macro in enumerate(macro_distribution):
                        data = response_data[macro_names[i]]
                        macro.protein = data["protein"]
                        macro.carbs = data["carbs"]
                        macro.fat = data["fat"]
                        macro.save(update_fields=["protein", "carbs", "fat"])

                    macros = {
                        "calorie": diet.calorie,
                        "balanced": MacroDistribution.objects.get(
                            plan_name="Balanced", user_diet=diet
                        ),
                        "lowcarbs": MacroDistribution.objects.get(
                            plan_name="Low Carb", user_diet=diet
                        ),
                        "lowfat": MacroDistribution.objects.get(
                            plan_name="Low Fat", user_diet=diet
                        ),
                        "highprotein": MacroDistribution.objects.get(
                            plan_name="High Protein", user_diet=diet
                        ),
                    }

                else:
                    stat_instance = form.save(commit=False)
                    stat_instance.user = user
                    stat_instance.save()

                    diet = Diet.objects.create(
                        calorie=response_data["calorie"],
                        stats=stat_instance,
                    )

                    macros = {
                        "calorie": response_data["calorie"],
                        "balanced": MacroDistribution.objects.create(
                            plan_name="Balanced", **balanced_data, user_diet=diet
                        ),
                        "lowcarbs": MacroDistribution.objects.create(
                            plan_name="Low Carb", **lowcarb_data, user_diet=diet
                        ),
                        "lowfat": MacroDistribution.objects.create(
                            plan_name="Low Fat", **lowfat_data, user_diet=diet
                        ),
                        "highprotein": MacroDistribution.objects.create(
                            plan_name="High Protein", **highprotein_data, user_diet=diet
                        ),
                    }

                return render(request, "calculator/profile.html", {"macros": macros})

            except requests.exceptions.RequestException as e:
                error_message = str(e)
                return render(
                    request, "calculator/error.html", {"error_message": error_message}
                )

    else:
        form = form_choice()
        context = {"form": form}

        return render(request, "calculator/home.html", context)


def metric(request):
    return calculate_macros(request, MetricForm)


def imperial(request):
    return calculate_macros(request, ImperialForm)


def landing_page(request):
    metric_form = MetricForm()
    imperial_form = ImperialForm()
    signup_form = SignUp()

    if request.user.is_authenticated:
        try:
            stat_instance = UserStat.objects.get(user=request.user)
            return redirect("profile", pk=stat_instance.pk)
        except UserStat.DoesNotExist:
            pass

    context = {
        "metric_form": metric_form,
        "imperial_form": imperial_form,
        "signup_form": signup_form,
    }

    return render(request, "calculator/home.html", context)


def profile(request, pk):
    user_stat = UserStat.objects.get(pk=pk)
    diet = Diet.objects.get(stats=user_stat)
    macros = MacroDistribution.objects.filter(user_diet=diet)
    macros = {
        "calorie": diet.calorie,
        "balanced": macros[0],
        "lowcarbs": macros[1],
        "lowfat": macros[2],
        "highprotein": macros[3],
    }
    return render(request, "calculator/profile.html", {"macros": macros})


def edit_profile(request, pk):
    current_record = UserStat.objects.get(pk=pk)
    metric_form = MetricForm(request.POST or None, instance=current_record)
    imperial_form = ImperialForm(request.POST or None, instance=current_record)

    return render(
        request,
        "calculator/edit_profile.html",
        {"metric_form": metric_form, "imperial_form": imperial_form},
    )
